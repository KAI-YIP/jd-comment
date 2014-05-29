#-*- coding:utf-8 -*  
import openpyxl.cell  
from openpyxl.reader.excel import load_workbook  
import MySQLdb  
import time  
  
#开始时间  
startTime = time.time()  
  
#读取excel2007文件  
wb = load_workbook(filename = r'/home/alber/data_base/tripadvisor/trip (1).xlsx')  
sheetnames = wb.get_sheet_names()  
ws = wb.get_sheet_by_name(sheetnames[1]) 
for rx in range(ws.get_highest_row()): 
                if (ws.cell(row=rx,column=4).value=="Hotel"):
	                print ws.cell(row=rx,column=11).value
                else:
	                pass


#显示有多少张表  
#print "Worksheet range(s):", wb.get_named_ranges()  
#print "Worksheet name(s):", wb.get_sheet_names()  

  
#取第一张表  
sheetnames = wb.get_sheet_names()  
ws = wb.get_sheet_by_name(sheetnames[0])  
  
#显示表名，表行数，表列数  
#print "Work Sheet Titile:",ws.title  
#print "Work Sheet Rows:",ws.get_highest_row()  
#print "Work Sheet Cols:",ws.get_highest_column()  

