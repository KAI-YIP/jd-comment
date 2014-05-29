#-*- coding:utf-8 -*
import os
from openpyxl.reader.excel import load_workbook
import MySQLdb
import time

# 开始时间
t1 = time.time()

# get name of file
namelist = os.listdir('/home/alber/data_base/tripadvisor')
# read xlsx
for name in namelist:
                    txtname=str(name[0:3]+".txt")
                    txtaddress=str("/home/alber/data_base/tripadvisor_result/"+txtname)
                    f=open(str(txtaddress),"a") 
                    name_address = str("/home/alber/data_base/tripadvisor/" + name)
                    print name_address
                    wb = load_workbook(filename=name_address)
                    sheetnames = wb.get_sheet_names()
                    ws = wb.get_sheet_by_name(sheetnames[1])
                    for rx in range(ws.get_highest_row()):
                                if (ws.cell(row=rx,column=4).value=="Hotel"):
                                                      content=ws.cell(row=rx,column=11).value
                                                      title=ws.cell(row=rx,column=8).value
                                                      if (content is not None):
                                                                result0="hotel-------"+str(title.encode('utf-8'))+" --------"+str(content.encode('utf-8')+'\n')                                                     
                                                                f.write(result0)
                                if (ws.cell(row=rx,column=4).value=="Restaurant"):
                                                    content=ws.cell(row=rx,column=11).value
                                                    title=ws.cell(row=rx,column=8).value
                                                    if (content is not None):
                                                               result1="restaurant-------"+str(title.encode('utf-8'))+" --------"+str(content.encode('utf-8')+'\n')  
                                                               f.write(result1)
                                if (ws.cell(row=rx,column=4).value=="Attraction"):
                                                     content=ws.cell(row=rx,column=11).value
                                                     title=ws.cell(row=rx,column=8).value
                                                     if (content is not None):
                                                               result2="attraction--------"+str(title.encode('utf-8'))+" --------"+str(content.encode('utf-8')+'\n')
                                                               f.write(result2)
                                else:
                                                   pass
f.close()

t2=time.time()
print("finished time："+str(t2-t1)+"秒。")



